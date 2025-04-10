import pandas as pd
import os
import sys
import logging
import traceback
import openpyxl
import numpy as np
from openpyxl.utils.exceptions import InvalidFileException

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

def diagnose_excel_file(file_path, specific_sheet=None):
    """Perform a comprehensive diagnosis of an Excel file with detailed error trapping"""
    logger.info(f"Advanced diagnosis of Excel file: {file_path}")

    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        return

    # Try different methods to read the Excel file
    try_methods = [
        {"name": "pandas default", "func": lambda: pd.ExcelFile(file_path)},
        {"name": "openpyxl", "func": lambda: pd.ExcelFile(file_path, engine='openpyxl')},
        {"name": "direct openpyxl", "func": lambda: openpyxl.load_workbook(file_path, read_only=True, data_only=True)}
    ]

    xl = None
    sheet_names = []
    successful_method = None

    # Try each method to open the file
    for method in try_methods:
        try:
            logger.info(f"Trying to open Excel file with {method['name']}...")
            if method['name'] == 'direct openpyxl':
                wb = method['func']()
                sheet_names = wb.sheetnames
                logger.info(f"Successfully opened file with {method['name']}")
                successful_method = method['name']
                break
            else:
                xl = method['func']()
                sheet_names = xl.sheet_names
                logger.info(f"Successfully opened file with {method['name']}")
                successful_method = method['name']
                break
        except Exception as e:
            logger.warning(f"Failed to open with {method['name']}: {str(e)}")
            logger.debug(traceback.format_exc())

    if not sheet_names:
        logger.error("Could not open the Excel file with any method")
        return

    logger.info(f"Found {len(sheet_names)} sheets: {', '.join(sheet_names)}")

    # If a specific sheet is requested, only examine that one
    if specific_sheet:
        if specific_sheet in sheet_names:
            sheets_to_examine = [specific_sheet]
        else:
            # Try case-insensitive matching
            sheet_name_lookup = {name.lower().strip(): name for name in sheet_names}
            if specific_sheet.lower().strip() in sheet_name_lookup:
                actual_sheet = sheet_name_lookup[specific_sheet.lower().strip()]
                logger.info(f"Using case-insensitive match: '{specific_sheet}' -> '{actual_sheet}'")
                sheets_to_examine = [actual_sheet]
            else:
                logger.error(f"Requested sheet '{specific_sheet}' not found in the Excel file")
                return
    else:
        sheets_to_examine = sheet_names

    # Examine each sheet
    for sheet_name in sheets_to_examine:
        logger.info(f"\n=== Examining sheet: {sheet_name} ===")

        # Try different methods to read the sheet
        sheet_methods = [
            {"name": "pandas default", "func": lambda: pd.read_excel(file_path, sheet_name=sheet_name)},
            {"name": "openpyxl engine", "func": lambda: pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')},
            {"name": "direct openpyxl", "func": lambda: get_sheet_data_openpyxl(file_path, sheet_name)}
        ]

        df = None
        sheet_successful_method = None

        for method in sheet_methods:
            try:
                logger.info(f"  Trying to read sheet with {method['name']}...")
                df = method['func']()
                logger.info(f"  Successfully read sheet with {method['name']}")
                sheet_successful_method = method['name']
                break
            except Exception as e:
                logger.warning(f"  Failed to read sheet with {method['name']}: {str(e)}")
                logger.debug(traceback.format_exc())

        if df is None:
            logger.error(f"  Could not read sheet '{sheet_name}' with any method")
            continue

        # Analyze the sheet
        try:
            logger.info(f"  Sheet has {len(df)} rows and {len(df.columns)} columns")

            # Check column names
            logger.info(f"  Column names:")
            for i, col in enumerate(df.columns):
                col_type = type(col).__name__
                col_repr = repr(col)
                col_str = str(col)
                logger.info(f"    {i+1}. {col_repr} (type: {col_type})")

                # Check for whitespace or special characters
                if isinstance(col, str):
                    if col != col.strip():
                        logger.warning(f"      Column has leading/trailing whitespace: '{col}' vs '{col.strip()}'")
                    if '\n' in col or '\r' in col or '\t' in col:
                        logger.warning(f"      Column contains special characters")

            # Check for duplicate column names
            if len(df.columns) != len(set(df.columns)):
                logger.warning("  Sheet contains duplicate column names!")
                col_counts = {}
                for col in df.columns:
                    col_counts[col] = col_counts.get(col, 0) + 1
                for col, count in col_counts.items():
                    if count > 1:
                        logger.warning(f"    Column '{col}' appears {count} times")

            # Check data types
            logger.info(f"  Column data types:")
            for col, dtype in df.dtypes.items():
                logger.info(f"    {col}: {dtype}")

            # Check for NaN values
            nan_counts = df.isna().sum()
            if nan_counts.sum() > 0:
                logger.info(f"  NaN value counts:")
                for col, count in nan_counts.items():
                    if count > 0:
                        logger.info(f"    {col}: {count} NaN values")

            # Check for problematic values
            logger.info(f"  Checking for problematic values...")
            for col in df.columns:
                try:
                    # Check for mixed data types
                    if df[col].dtype == 'object':
                        types = df[col].apply(lambda x: type(x).__name__ if not pd.isna(x) else 'nan').unique()
                        if len(types) > 1 and 'nan' in types:
                            types = [t for t in types if t != 'nan']
                        if len(types) > 1:
                            logger.warning(f"    Column '{col}' has mixed data types: {', '.join(types)}")

                    # Check for extremely long values
                    if df[col].dtype == 'object':
                        max_len = df[col].astype(str).apply(len).max()
                        if max_len > 1000:
                            logger.warning(f"    Column '{col}' has very long values (max length: {max_len})")
                except Exception as e:
                    logger.warning(f"    Error checking column '{col}': {str(e)}")

            # Sample data
            logger.info(f"  Sample data (first 2 rows):")
            try:
                sample = df.head(2)
                for i, row in sample.iterrows():
                    logger.info(f"    Row {i+1}:")
                    for col in sample.columns:
                        val = row[col]
                        val_type = type(val).__name__
                        val_repr = repr(val)
                        logger.info(f"      {col}: {val_repr} (type: {val_type})")
            except Exception as e:
                logger.warning(f"    Error displaying sample data: {str(e)}")

        except Exception as e:
            logger.error(f"  Error analyzing sheet '{sheet_name}': {str(e)}")
            logger.error(traceback.format_exc())

def get_sheet_data_openpyxl(file_path, sheet_name):
    """Get sheet data using openpyxl directly"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name]

    # Get all data from the worksheet
    data = []
    for row in ws.rows:
        data.append([cell.value for cell in row])

    # Convert to DataFrame
    if data:
        df = pd.DataFrame(data[1:], columns=data[0])
        return df
    else:
        return pd.DataFrame()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python advanced_excel_diagnostic.py <excel_file_path> [sheet_name]")
        sys.exit(1)

    file_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

    diagnose_excel_file(file_path, sheet_name)
