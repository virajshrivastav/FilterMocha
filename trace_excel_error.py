import pandas as pd
import os
import sys
import logging
import traceback
import openpyxl
import json
import time
from flask import Flask, request, jsonify
from excel_standardizer_improved import ExcelStandardizer

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,  # Set to DEBUG for maximum detail
    format='%(asctime)s - %(levelname)s - %(name)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('excel_trace.log')  # Also log to a file
    ]
)
logger = logging.getLogger('excel_tracer')

# Create a small Flask app to simulate the API endpoint
app = Flask(__name__)

@app.route('/analyze_file', methods=['POST'])
def analyze_file_endpoint():
    """Endpoint to analyze an Excel file with detailed error tracing"""
    try:
        # Get file path and sheet name from request
        data = request.json
        file_path = data.get('file_path')
        sheet_name = data.get('sheet_name')
        
        logger.info(f"Analyzing file: {file_path}, sheet: {sheet_name}")
        
        # Create a standardizer instance
        standardizer = ExcelStandardizer()
        
        # Add a hook to trace the execution
        def trace_execution(frame, event, arg):
            if event == 'line':
                filename = frame.f_code.co_filename
                lineno = frame.f_lineno
                function = frame.f_code.co_name
                
                # Only trace our code
                if 'excel_standardizer_improved.py' in filename:
                    # Get local variables
                    local_vars = {}
                    for var_name, var_val in frame.f_locals.items():
                        try:
                            # Try to convert to string, but limit size
                            if isinstance(var_val, (str, int, float, bool, type(None))):
                                local_vars[var_name] = str(var_val)
                            elif isinstance(var_val, (list, tuple)):
                                local_vars[var_name] = f"[{type(var_val).__name__}] length: {len(var_val)}"
                            elif isinstance(var_val, dict):
                                local_vars[var_name] = f"[dict] keys: {list(var_val.keys())[:5]}"
                            elif isinstance(var_val, pd.DataFrame):
                                local_vars[var_name] = f"[DataFrame] shape: {var_val.shape}"
                            else:
                                local_vars[var_name] = f"[{type(var_val).__name__}]"
                        except:
                            local_vars[var_name] = "[Error getting value]"
                    
                    logger.debug(f"TRACE: {filename}:{lineno} - {function} - Locals: {json.dumps(local_vars, indent=2)}")
            
            return trace_execution
        
        # Enable tracing
        sys.settrace(trace_execution)
        
        try:
            # Try to analyze the file
            start_time = time.time()
            result = standardizer.analyze_file(file_path, sheet_name)
            end_time = time.time()
            
            # Disable tracing
            sys.settrace(None)
            
            # Format the result
            columns_info, file_shape, sheet_names, selected_sheet, sheet_info = result
            
            response = {
                'success': True,
                'message': 'File analyzed successfully',
                'data': {
                    'columns': len(columns_info),
                    'rows': file_shape[0] if file_shape else 0,
                    'sheets': sheet_names,
                    'selected_sheet': selected_sheet,
                    'execution_time': f"{end_time - start_time:.2f} seconds"
                }
            }
            
            logger.info(f"Analysis successful: {response}")
            return jsonify(response)
            
        except Exception as e:
            # Disable tracing
            sys.settrace(None)
            
            # Log the full traceback
            logger.error(f"Error analyzing file: {str(e)}")
            logger.error(traceback.format_exc())
            
            response = {
                'success': False,
                'message': f"Error analyzing file: {str(e)}",
                'traceback': traceback.format_exc()
            }
            
            return jsonify(response), 500
    
    except Exception as e:
        logger.error(f"Endpoint error: {str(e)}")
        logger.error(traceback.format_exc())
        
        return jsonify({
            'success': False,
            'message': f"Endpoint error: {str(e)}",
            'traceback': traceback.format_exc()
        }), 500

def trace_file(file_path, sheet_name=None):
    """Directly trace file analysis without using the Flask app"""
    try:
        logger.info(f"Directly tracing file: {file_path}, sheet: {sheet_name}")
        
        # Create a standardizer instance
        standardizer = ExcelStandardizer()
        
        # Try to analyze the file
        try:
            # First check if the file exists
            if not os.path.exists(file_path):
                logger.error(f"File not found: {file_path}")
                return
            
            # Try to open the file with different methods
            try:
                # Try with pandas
                logger.info("Trying to open with pandas...")
                xl = pd.ExcelFile(file_path)
                sheet_names = xl.sheet_names
                logger.info(f"Successfully opened with pandas. Sheets: {sheet_names}")
            except Exception as e:
                logger.error(f"Failed to open with pandas: {str(e)}")
                
                try:
                    # Try with openpyxl
                    logger.info("Trying to open with openpyxl...")
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    sheet_names = wb.sheetnames
                    logger.info(f"Successfully opened with openpyxl. Sheets: {sheet_names}")
                except Exception as e:
                    logger.error(f"Failed to open with openpyxl: {str(e)}")
                    logger.error("Could not open the file with any method")
                    return
            
            # If a specific sheet is requested, check if it exists
            if sheet_name:
                # Try case-insensitive matching
                sheet_name_lookup = {name.lower().strip(): name for name in sheet_names}
                
                if sheet_name in sheet_names:
                    actual_sheet = sheet_name
                    logger.info(f"Sheet '{sheet_name}' found (exact match)")
                elif sheet_name.lower().strip() in sheet_name_lookup:
                    actual_sheet = sheet_name_lookup[sheet_name.lower().strip()]
                    logger.info(f"Sheet '{sheet_name}' found (case-insensitive match): '{actual_sheet}'")
                else:
                    logger.error(f"Sheet '{sheet_name}' not found in file. Available sheets: {sheet_names}")
                    return
            else:
                # If no sheet specified, use the first one
                actual_sheet = sheet_names[0]
                logger.info(f"No sheet specified, using first sheet: '{actual_sheet}'")
            
            # Try to read the sheet with different methods
            logger.info(f"Attempting to read sheet '{actual_sheet}'...")
            
            # Method 1: pandas default
            try:
                logger.info("Method 1: Using pandas default...")
                df = pd.read_excel(file_path, sheet_name=actual_sheet)
                logger.info(f"Success! Read {len(df)} rows and {len(df.columns)} columns")
                logger.info(f"Column names: {df.columns.tolist()}")
                logger.info(f"Data types: {df.dtypes.to_dict()}")
                
                # Check for problematic values
                for col in df.columns:
                    try:
                        # Check for mixed data types
                        if df[col].dtype == 'object':
                            types = df[col].apply(lambda x: type(x).__name__ if not pd.isna(x) else 'nan').unique()
                            if len(types) > 1 and 'nan' in types:
                                types = [t for t in types if t != 'nan']
                            if len(types) > 1:
                                logger.warning(f"Column '{col}' has mixed data types: {', '.join(types)}")
                        
                        # Check for NaN values
                        nan_count = df[col].isna().sum()
                        if nan_count > 0:
                            logger.info(f"Column '{col}' has {nan_count} NaN values")
                        
                        # Check for extremely long values
                        if df[col].dtype == 'object':
                            max_len = df[col].astype(str).apply(len).max()
                            if max_len > 1000:
                                logger.warning(f"Column '{col}' has very long values (max length: {max_len})")
                    except Exception as e:
                        logger.warning(f"Error checking column '{col}': {str(e)}")
            except Exception as e:
                logger.error(f"Method 1 failed: {str(e)}")
                
                # Method 2: openpyxl engine
                try:
                    logger.info("Method 2: Using openpyxl engine...")
                    df = pd.read_excel(file_path, sheet_name=actual_sheet, engine='openpyxl')
                    logger.info(f"Success! Read {len(df)} rows and {len(df.columns)} columns")
                except Exception as e:
                    logger.error(f"Method 2 failed: {str(e)}")
                    
                    # Method 3: converters
                    try:
                        logger.info("Method 3: Using converters...")
                        converters = {i: str for i in range(100)}
                        df = pd.read_excel(file_path, sheet_name=actual_sheet, converters=converters)
                        logger.info(f"Success! Read {len(df)} rows and {len(df.columns)} columns")
                    except Exception as e:
                        logger.error(f"Method 3 failed: {str(e)}")
                        
                        # Method 4: direct openpyxl
                        try:
                            logger.info("Method 4: Using direct openpyxl...")
                            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                            ws = wb[actual_sheet]
                            
                            # Get all data from the worksheet
                            data = []
                            for row in ws.rows:
                                data.append([cell.value for cell in row])
                            
                            if data:
                                df = pd.DataFrame(data[1:], columns=data[0])
                                logger.info(f"Success! Read {len(df)} rows and {len(df.columns)} columns")
                            else:
                                logger.error("No data found in sheet")
                        except Exception as e:
                            logger.error(f"Method 4 failed: {str(e)}")
                            logger.error("All methods failed to read the sheet")
                            return
            
            # Now try to analyze the file with the standardizer
            logger.info("Now trying to analyze with the standardizer...")
            result = standardizer.analyze_file(file_path, sheet_name)
            
            # Format the result
            columns_info, file_shape, sheet_names, selected_sheet, sheet_info = result
            
            logger.info(f"Analysis successful!")
            logger.info(f"Columns: {len(columns_info)}")
            logger.info(f"Rows: {file_shape[0] if file_shape else 0}")
            logger.info(f"Selected sheet: {selected_sheet}")
            
        except Exception as e:
            # Log the full traceback
            logger.error(f"Error analyzing file: {str(e)}")
            logger.error(traceback.format_exc())
    
    except Exception as e:
        logger.error(f"Tracing error: {str(e)}")
        logger.error(traceback.format_exc())

if __name__ == "__main__":
    if len(sys.argv) > 1:
        # Direct tracing mode
        file_path = sys.argv[1]
        sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
        trace_file(file_path, sheet_name)
    else:
        # Start the Flask app
        app.run(host='127.0.0.1', port=5052, debug=True)
